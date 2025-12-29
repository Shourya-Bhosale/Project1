# Standard library imports
import os
import traceback

# Django imports
from django.conf import settings
from django.db.models.signals import post_migrate, post_save
from django.dispatch import receiver


# Third-party imports (lazy imports in functions to avoid dependency issues)
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import Font, PatternFill, Alignment


# Helper Functions
def export_order_to_excel(order):
    """Export order to Excel file. Creates file if it doesn't exist, appends if it does."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        excel_file_path = os.path.join(settings.BASE_DIR, 'orders.xlsx')
        
        if os.path.exists(excel_file_path):
            wb = load_workbook(excel_file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Orders"
            
            headers = [
                'Order Number', 'Date', 'Customer Name', 'Email', 'Phone',
                'Address Line 1', 'Address Line 2', 'City', 'State', 'Postal Code',
                'Latitude', 'Longitude', 'Payment Method', 'Payment Status',
                'Payment Reference', 'Total Amount', 'Items', 'Notes',
                'Razorpay Order ID', 'Razorpay Payment ID'
            ]
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.freeze_panes = 'A2'
        
        row_num = ws.max_row + 1
        
        items_list = []
        for item in order.items.all():
            items_list.append(f"{item.product.name} x{item.quantity} @ ₹{item.unit_price}")
        items_str = "; ".join(items_list)
        
        ws.cell(row=row_num, column=1, value=order.order_number)
        ws.cell(row=row_num, column=2, value=order.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_num, column=3, value=order.customer_name)
        ws.cell(row=row_num, column=4, value=order.email)
        ws.cell(row=row_num, column=5, value=order.phone)
        ws.cell(row=row_num, column=6, value=order.address_line1)
        ws.cell(row=row_num, column=7, value=order.address_line2 or '')
        ws.cell(row=row_num, column=8, value=order.city)
        ws.cell(row=row_num, column=9, value=order.state)
        ws.cell(row=row_num, column=10, value=order.postal_code)
        ws.cell(row=row_num, column=11, value=str(order.latitude) if order.latitude else '')
        ws.cell(row=row_num, column=12, value=str(order.longitude) if order.longitude else '')
        ws.cell(row=row_num, column=13, value=order.get_payment_method_display())
        ws.cell(row=row_num, column=14, value=order.payment_status)
        ws.cell(row=row_num, column=15, value=order.payment_reference or '')
        ws.cell(row=row_num, column=16, value=order.total_amount)
        ws.cell(row=row_num, column=17, value=items_str)
        ws.cell(row=row_num, column=18, value=order.notes or '')
        ws.cell(row=row_num, column=19, value=order.razorpay_order_id or '')
        ws.cell(row=row_num, column=20, value=order.razorpay_payment_id or '')
        
        for col in range(1, 21):
            cell = ws.cell(row=row_num, column=col)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 20
        ws.column_dimensions['P'].width = 12
        ws.column_dimensions['Q'].width = 40
        ws.column_dimensions['R'].width = 30
        ws.column_dimensions['S'].width = 25
        ws.column_dimensions['T'].width = 25
        
        wb.save(excel_file_path)
        print(f"[EXCEL] Order #{order.order_number} exported to {excel_file_path}")
        return True
    except Exception as e:
        print(f"[EXCEL ERROR] Failed to export order to Excel: {str(e)}")
        traceback.print_exc()
        return False


# Signal Handlers
@receiver(post_save)
def save_order_to_excel(sender, instance, created, **kwargs):
    """Automatically export order to Excel when saved."""
    try:
        from .models import Order
        if isinstance(instance, Order) and created and instance.order_number:
            export_order_to_excel(instance)
    except Exception as e:
        print(f"[EXCEL] Error exporting order to Excel: {str(e)}")


@receiver(post_migrate)
def seed_products(sender, **kwargs):
    """Seed initial products after migrations."""
    try:
        from .models import Product
        if Product.objects.count() == 0:
            Product.objects.bulk_create(
                [
                    Product(
                        name='Gir Cow Ghee 1L',
                        size_ml=1000,
                        price=1999,
                        description='Pure A2 Gir Cow Bilona Ghee (1L).',
                        image_url='/static/store/images/product_1l.jpg',
                    ),
                    Product(
                        name='Gir Cow Ghee 500ml',
                        size_ml=500,
                        price=1035,
                        description='Pure A2 Gir Cow Bilona Ghee (500ml).',
                        image_url='/static/store/images/product_500ml.jpg',
                    ),
                    Product(
                        name='Gir Cow Ghee 250ml',
                        size_ml=250,
                        price=549,
                        description='Pure A2 Gir Cow Bilona Ghee (250ml).',
                        image_url='/static/store/images/product_250ml.jpg',
                    ),
                ]
            )
    except Exception:
        pass
